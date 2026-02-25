"""
Export microbiome annotation JSONs to Excel — one tab per paper + summary tab.
Usage: python json_to_excel.py <input_dir_or_file> <output.xlsx>

  Directory of per-paper JSONs:  python json_to_excel.py ./papers_for_annotation_validation/ annotations_review.xlsx
  Single merged JSON:             python json_to_excel.py merged.json annotations_review.xlsx
"""

import json
import sys
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection

# ---------------------------------------------------------------------------
# I/O
# ---------------------------------------------------------------------------

def load_papers(path):
    if os.path.isdir(path):
        files = sorted(glob.glob(os.path.join(path, "paper_*.json")))
        if not files:
            raise FileNotFoundError(f"No paper_*.json files found in {path}")
        papers = []
        for f in files:
            with open(f) as fh:
                papers.append(json.load(fh))
        print(f"Loaded {len(papers)} paper JSONs from {path}")
        return papers
    else:
        with open(path) as f:
            data = json.load(f)
        papers = []
        for paper_id, paper in data.items():
            paper.setdefault("paper_id", paper_id)
            paper.setdefault("paper_title", paper.get("title", ""))
            papers.append(paper)
        return papers

def flatten_paper(paper):
    paper_id = paper.get("paper_id", "")
    title    = paper.get("paper_title") or paper.get("title", "")
    rows = []
    for rel in paper.get("relationships", []):
        meta = rel.get("_annotation_metadata", {})
        rows.append({
            "taxon_name":        rel.get("taxon_name", ""),
            "domain":            rel.get("domain", ""),
            "taxonomic_level":   rel.get("taxonomic_level", ""),
            "primary_disease":   rel.get("primary_disease", ""),
            "direction":         rel.get("direction", ""),
            "change_context":    rel.get("change_context", ""),
            "sample_site":       rel.get("sample_site", ""),
            "p_value":           "" if rel.get("p_value") is None else str(rel.get("p_value")),
            "confidence":        rel.get("confidence", ""),
            "agreement":         meta.get("agreement_level", ""),
            "queen_decision":    meta.get("queen_decision", ""),
            "verbatim_evidence": meta.get("verbatim_evidence", ""),
            "status":            "",
            "correction":        "",
            "sam_notes":         "",
        })
    return paper_id, title, rows

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

# (col_name, width, is_review, is_transparency)
COLUMNS = [
    ("taxon_name",        24,  False, False),
    ("domain",            12,  False, False),
    ("taxonomic_level",   16,  False, False),
    ("primary_disease",   24,  False, False),
    ("direction",         12,  False, False),
    ("change_context",    20,  False, False),
    ("sample_site",       13,  False, False),
    ("p_value",           10,  False, False),
    ("confidence",        12,  False, False),
    ("agreement",         16,  False, True),
    ("queen_decision",    16,  False, True),
    ("verbatim_evidence", 48,  False, True),
    ("status",            14,  True,  False),
    ("correction",        25,  True,  False),
    ("sam_notes",         35,  True,  False),
]

DIRECTION_COLORS = {
    "increased": "C6EFCE",
    "decreased": "FFC7CE",
    "unchanged": "FFEB9C",
    "unclear":   "D9D9D9",
}

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def mfont(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)

def mfill(hex_color):
    return PatternFill("solid", start_color=hex_color)

# ---------------------------------------------------------------------------
# Per-paper sheet
# ---------------------------------------------------------------------------

def write_paper_sheet(wb, paper_id, title, rows):
    # Sheet name: "P141" style — Excel tab names max 31 chars
    sheet_name = f"P{paper_id}"[:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.protection.sheet   = True
    ws.protection.password = ""

    col_names = [c[0] for c in COLUMNS]

    # Title banner row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    banner = ws.cell(row=1, column=1, value=f"Paper {paper_id} — {title}")
    banner.font      = mfont(bold=True, color="FFFFFF", size=11)
    banner.fill      = mfill("1F4E79")
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Header row
    for ci, (col_name, width, is_review, is_trans) in enumerate(COLUMNS, start=1):
        if is_review:
            hdr_color = "375623"
        elif is_trans:
            hdr_color = "4A4A6A"
        else:
            hdr_color = "1F4E79"
        cell = ws.cell(row=2, column=ci, value=col_name.replace("_", " ").upper())
        cell.font      = mfont(bold=True, color="FFFFFF")
        cell.fill      = mfill(hdr_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    if rows:
        for ri, row in enumerate(rows, start=3):
            alt = ri % 2 == 0
            for ci, (col_name, _, is_review, is_trans) in enumerate(COLUMNS, start=1):
                value = row.get(col_name, "")
                cell  = ws.cell(row=ri, column=ci, value=value)
                cell.font   = mfont()
                cell.border = BORDER
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=(col_name in {"verbatim_evidence", "sam_notes", "correction"})
                )
                if is_review:
                    cell.fill       = mfill("EAF4E8")
                    cell.protection = Protection(locked=False)
                elif is_trans:
                    cell.fill       = mfill("EEEEF5" if alt else "E8E8F0")
                    cell.protection = Protection(locked=True)
                elif col_name == "direction" and value in DIRECTION_COLORS:
                    cell.fill       = mfill(DIRECTION_COLORS[value])
                    cell.protection = Protection(locked=True)
                else:
                    cell.fill       = mfill("EBF3FB" if alt else "FFFFFF")
                    cell.protection = Protection(locked=True)
            ws.row_dimensions[ri].height = 20

        # Dropdown for status
        status_ci  = col_names.index("status") + 1
        status_col = get_column_letter(status_ci)
        last_row   = len(rows) + 2
        dv = DataValidation(
            type="list",
            formula1='"approved,rejected,modified"',
            allow_blank=True,
            showDropDown=False
        )
        dv.sqref = f"{status_col}3:{status_col}{last_row}"
        ws.add_data_validation(dv)

    else:
        # Empty paper — leave blank rows for Sam to fill in manually
        msg = ws.cell(row=3, column=1,
                      value="No relationships extracted — please review paper manually and add below if needed.")
        msg.font      = mfont(bold=True, color="C00000")
        msg.fill      = mfill("FFF2CC")
        msg.alignment = Alignment(horizontal="left", vertical="center")
        msg.protection = Protection(locked=False)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(COLUMNS))

        # Blank editable rows for manual entry
        for ri in range(4, 14):
            for ci, (col_name, _, is_review, is_trans) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=ri, column=ci, value="")
                cell.border     = BORDER
                cell.font       = mfont()
                cell.protection = Protection(locked=False)
                if is_review:
                    cell.fill = mfill("EAF4E8")
                elif is_trans:
                    cell.fill = mfill("EEEEF5")
                else:
                    cell.fill = mfill("FFFFFF")
            ws.row_dimensions[ri].height = 20

    return sheet_name

# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def write_summary_sheet(wb, paper_meta, sheet_names):
    """
    paper_meta: list of (paper_id, title, n_rels, sheet_name)
    """
    ws = wb["Summary"]

    # Header banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    banner = ws.cell(row=1, column=1, value="Annotation Review — Summary")
    banner.font      = mfont(bold=True, color="FFFFFF", size=12)
    banner.fill      = mfill("1F4E79")
    banner.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["Paper ID", "Title", "Extracted", "Approved", "Rejected", "Modified", "Pending", "Sheet"]
    col_widths = [10, 50, 11, 11, 11, 11, 11, 10]
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = mfont(bold=True, color="FFFFFF")
        cell.fill      = mfill("2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    total_row = len(paper_meta) + 3  # where totals go

    for ri, (paper_id, title, n_rels, sheet_name) in enumerate(paper_meta, start=3):
        alt  = ri % 2 == 0
        fill = mfill("EBF3FB" if alt else "FFFFFF")

        # Paper ID — hyperlink to sheet
        ws.cell(row=ri, column=1, value=paper_id).font   = mfont()
        ws.cell(row=ri, column=1).fill   = fill
        ws.cell(row=ri, column=1).border = BORDER
        ws.cell(row=ri, column=1).alignment = Alignment(horizontal="center")

        # Title
        ws.cell(row=ri, column=2, value=title).font   = mfont()
        ws.cell(row=ri, column=2).fill   = fill
        ws.cell(row=ri, column=2).border = BORDER

        # Extracted count
        ws.cell(row=ri, column=3, value=n_rels).font   = mfont()
        ws.cell(row=ri, column=3).fill   = fill
        ws.cell(row=ri, column=3).border = BORDER
        ws.cell(row=ri, column=3).alignment = Alignment(horizontal="center")

        # Approved / Rejected / Modified — formulas referencing paper sheet
        if n_rels > 0:
            status_col_letter = get_column_letter(COLUMNS.index(next(c for c in COLUMNS if c[0] == "status")) + 1)
            last_data_row = n_rels + 2
            ref = f"'{sheet_name}'!{status_col_letter}3:{status_col_letter}{last_data_row}"
            ws.cell(row=ri, column=4, value=f'=COUNTIF({ref},"approved")')
            ws.cell(row=ri, column=5, value=f'=COUNTIF({ref},"rejected")')
            ws.cell(row=ri, column=6, value=f'=COUNTIF({ref},"modified")')
            ws.cell(row=ri, column=7, value=f'={n_rels}-D{ri}-E{ri}-F{ri}')
        else:
            for col in [4, 5, 6]:
                ws.cell(row=ri, column=col, value=0)
            ws.cell(row=ri, column=7, value="manual")

        for col in range(4, 8):
            ws.cell(row=ri, column=col).font      = mfont()
            ws.cell(row=ri, column=col).fill      = fill
            ws.cell(row=ri, column=col).border    = BORDER
            ws.cell(row=ri, column=col).alignment = Alignment(horizontal="center")

        # Sheet link
        link_cell = ws.cell(row=ri, column=8, value=f"→ {sheet_name}")
        link_cell.font      = Font(color="0563C1", underline="single", name="Arial", size=10)
        link_cell.fill      = fill
        link_cell.border    = BORDER
        link_cell.alignment = Alignment(horizontal="center")
        link_cell.hyperlink = f"#'{sheet_name}'!A1"

    # Totals row
    tr = total_row
    ws.cell(row=tr, column=1, value="TOTAL").font      = mfont(bold=True)
    ws.cell(row=tr, column=1).fill   = mfill("D6E4F0")
    ws.cell(row=tr, column=1).border = BORDER
    ws.cell(row=tr, column=2).fill   = mfill("D6E4F0")
    ws.cell(row=tr, column=2).border = BORDER
    for col in range(3, 8):
        ws.cell(row=tr, column=col, value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr-1})")
        ws.cell(row=tr, column=col).font      = mfont(bold=True)
        ws.cell(row=tr, column=col).fill      = mfill("D6E4F0")
        ws.cell(row=tr, column=col).border    = BORDER
        ws.cell(row=tr, column=col).alignment = Alignment(horizontal="center")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def write_excel(papers, out_path):
    wb = Workbook()
    # Rename default sheet to Summary
    wb.active.title = "Summary"

    paper_meta  = []
    sheet_names = []

    for paper in papers:
        paper_id, title, rows = flatten_paper(paper)
        sheet_name = write_paper_sheet(wb, paper_id, title, rows)
        paper_meta.append((paper_id, title, len(rows), sheet_name))
        sheet_names.append(sheet_name)

    write_summary_sheet(wb, paper_meta, sheet_names)

    # Move Summary to front
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    wb.save(out_path)
    total_rels   = sum(m[2] for m in paper_meta)
    empty_papers = sum(1 for m in paper_meta if m[2] == 0)
    print(f"Saved {total_rels} relationships from {len(papers)} papers ({empty_papers} empty) -> {out_path}")

def main():
    in_path  = sys.argv[1] if len(sys.argv) > 1 else "."
    out_path = sys.argv[2] if len(sys.argv) > 2 else "annotations_review.xlsx"
    papers   = load_papers(in_path)
    write_excel(papers, out_path)

if __name__ == "__main__":
    main()